from flask import Flask, request, send_file, render_template_string, abort
import pandas as pd
from io import BytesIO
import openpyxl

app = Flask(__name__)

HTML_TEMPLATE = """
<!doctype html>
<html lang=\"ko\">
  <head>
    <meta charset=\"utf-8\">
    <title>운송장 번호 병합기</title>
  </head>
  <body style=\"font-family: sans-serif; text-align: center; margin-top: 50px;\">
    <h1>📦 운송장번호 자동 병합 도구</h1>
    <form action=\"/merge\" method=\"post\" enctype=\"multipart/form-data\">
      <label>📄 배송리스트 파일 (.xlsx):</label><br>
      <input type=\"file\" name=\"delivery_file\" accept=\".xlsx\" required><br><br>
      <label>📄 운송장번호 포함 파일 (.xlsx):</label><br>
      <input type=\"file\" name=\"tracking_file\" accept=\".xlsx\" required><br><br>
      <button type=\"submit\" style=\"font-size: 16px;\">운송장 병합하기</button>
    </form>
  </body>
</html>
"""

@app.route("/", methods=["GET"])
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route("/merge", methods=["POST"])
def merge():
    delivery_file = request.files['delivery_file']
    tracking_file = request.files['tracking_file']

    delivery_df = pd.read_excel(delivery_file)
    tracking_wb = openpyxl.load_workbook(tracking_file, data_only=True)

    found = False
    for sheetname in tracking_wb.sheetnames:
        ws = tracking_wb[sheetname]
        headers = [cell.value for cell in ws[1]]
        if "상품주문번호" in headers and "등기번호" in headers:
            주문번호_idx = headers.index("상품주문번호")
            운송장_idx = headers.index("등기번호")
            found = True
            break

    if not found:
        return "❌ 업로드한 운송장 파일에서 '상품주문번호'와 '등기번호' 열을 찾을 수 없습니다."

    tracking_dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        주문번호 = str(row[주문번호_idx]).strip() if row[주문번호_idx] else None
        운송장번호 = str(row[운송장_idx]).strip() if row[운송장_idx] else None
        if 주문번호:
            tracking_dict[주문번호] = 운송장번호

    delivery_df["운송장번호"] = delivery_df["주문번호"].astype(str).map(tracking_dict)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        delivery_df.to_excel(writer, index=False, sheet_name="배송리스트")

    output.seek(0)
    return send_file(output, as_attachment=True,
                     download_name="배송리스트_운송장포함.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == "__main__":
    app.run(debug=True)
